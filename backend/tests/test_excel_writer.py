from io import BytesIO

from openpyxl import load_workbook

from app.schemas import LineItem, ParseResult
from app.services.excel_writer import PO_ITEMS_HEADERS, SUMMARY_HEADERS, build_workbook


def _sample_result() -> ParseResult:
    return ParseResult(
        source_file="4515457833 WHIRLPOOL.PDF",
        po_number="4515457833",
        po_date="02/17/2026",
        ship_to="01RI\nJabil Circuit India Pvt Ltd (EOU Unit)",
        status="parsed",
        line_items=[
            LineItem(
                item="1",
                material="W10165202",
                description="TRANSISTOR PNP 100MA 45V HFE150 SOT23",
                manufacturer_part_number="LBC857BLT1G",
                uom="EA",
                total_qty="78,000",
                qty_recd="0",
                qty_retd="0",
                unit_price="0.",
                item_value="0.00",
                due_date="10/03/2026",
            )
        ],
    )


def _sample_result_with_item_warning() -> ParseResult:
    return ParseResult(
        source_file="4515457833 WHIRLPOOL.PDF",
        po_number="4515457833",
        po_date="02/17/2026",
        ship_to="01RI\nJabil Circuit India Pvt Ltd (EOU Unit)",
        status="parsed",
        warnings=["Item 1: Description not found"],
        line_items=[
            LineItem(
                item="1",
                material="W10165202",
                description="TRANSISTOR PNP 100MA 45V HFE150 SOT23",
                manufacturer_part_number="LBC857BLT1G",
                uom="EA",
                total_qty="78,000",
                qty_recd="0",
                qty_retd="0",
                unit_price="0.",
                item_value="0.00",
                due_date="10/03/2026",
                warnings=["Description not found"],
            ),
            LineItem(
                item="2",
                material="W10165203",
                description="TRANSISTOR PNP 100MA 45V HFE150 SOT23",
                uom="EA",
                total_qty="12,000",
                qty_recd="0",
                qty_retd="0",
                unit_price="0.",
                item_value="0.00",
                due_date="10/03/2026",
            ),
        ],
    )


def test_build_workbook_creates_fixed_sheets_and_headers():
    workbook_bytes = build_workbook([_sample_result()])

    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == ["PO Items", "Parse Summary"]
    po_items_headers = [cell.value for cell in workbook["PO Items"][1]]
    assert po_items_headers == PO_ITEMS_HEADERS
    assert po_items_headers[7] == "Manufacturer Part Number"
    assert [cell.value for cell in workbook["Parse Summary"][1]] == SUMMARY_HEADERS


def test_build_workbook_writes_one_row_per_line_item():
    workbook = load_workbook(BytesIO(build_workbook([_sample_result()])))
    row = [cell.value for cell in workbook["PO Items"][2]]

    assert row[0] == "4515457833 WHIRLPOOL.PDF"
    assert row[1] == "4515457833"
    assert row[2] == "02/17/2026"
    assert row[5] == "W10165202"
    assert row[6] == "TRANSISTOR PNP 100MA 45V HFE150 SOT23"
    assert row[7] == "LBC857BLT1G"
    assert row[14] == "10/03/2026"
    assert row[15] == "parsed"


def test_build_workbook_writes_summary_for_failed_file():
    failed = ParseResult(source_file="bad.pdf", status="failed", error="PO number/date not found")

    workbook = load_workbook(BytesIO(build_workbook([failed])))
    row = [cell.value for cell in workbook["Parse Summary"][2]]

    assert row[0] == "bad.pdf"
    assert row[3] == "failed"
    assert row[4] == 0
    assert row[5] == 0
    assert row[7] == "PO number/date not found"


def test_build_workbook_keeps_item_warnings_on_correct_rows_and_dedupes_summary():
    workbook = load_workbook(BytesIO(build_workbook([_sample_result_with_item_warning()])))

    item_one = [cell.value for cell in workbook["PO Items"][2]]
    item_two = [cell.value for cell in workbook["PO Items"][3]]
    summary = [cell.value for cell in workbook["Parse Summary"][2]]

    assert item_one[15] == "parsed; Description not found"
    assert item_two[15] == "parsed"
    assert summary[5] == 1
    assert summary[6] == "Item 1: Description not found"
