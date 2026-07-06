from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import create_app

SAMPLE = Path("/home/eureka/catch/4515457833 WHIRLPOOL.PDF")


def _client() -> TestClient:
    return TestClient(create_app())


def _auth_headers(client: TestClient) -> dict[str, str]:
    response = client.post(
        "/api/auth/login",
        json={"username": "buyer", "password": "secret-password"},
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def test_parse_endpoint_requires_authentication():
    client = _client()

    response = client.post("/api/files/parse", files=[])

    assert response.status_code in {401, 403}


def test_parse_endpoint_extracts_uploaded_pdf():
    client = _client()
    headers = _auth_headers(client)

    with SAMPLE.open("rb") as file_obj:
        response = client.post(
            "/api/files/parse",
            headers=headers,
            files=[("files", (SAMPLE.name, file_obj, "application/pdf"))],
        )

    assert response.status_code == 200
    body = response.json()
    assert len(body["results"]) == 1
    result = body["results"][0]
    assert result["source_file"] == SAMPLE.name
    assert result["po_number"] == "4515457833"
    assert result["line_items"][0]["material"] == "W10165202"


def test_parse_endpoint_rejects_non_pdf_upload():
    client = _client()
    headers = _auth_headers(client)

    response = client.post(
        "/api/files/parse",
        headers=headers,
        files=[("files", ("notes.txt", b"hello", "text/plain"))],
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Only PDF files are supported"


def test_export_endpoint_returns_excel_file():
    client = _client()
    headers = _auth_headers(client)
    payload = {
        "results": [
            {
                "source_file": SAMPLE.name,
                "po_number": "4515457833",
                "po_date": "02/17/2026",
                "ship_to": "01RI\nJabil Circuit India Pvt Ltd (EOU Unit)",
                "status": "parsed",
                "warnings": [],
                "error": None,
                "line_items": [
                    {
                        "item": "1",
                        "material": "W10165202",
                        "description": "TRANSISTOR PNP 100MA 45V HFE150 SOT23",
                        "uom": "EA",
                        "total_qty": "78,000",
                        "qty_recd": "0",
                        "qty_retd": "0",
                        "unit_price": "0.",
                        "item_value": "0.00",
                        "due_date": "10/03/2026",
                        "warnings": [],
                    }
                ],
            }
        ]
    }

    response = client.post("/api/excel/export", headers=headers, json=payload)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "purchase_orders.xlsx" in response.headers["content-disposition"]

    workbook_path = Path("/tmp/purchase_orders_api_test.xlsx")
    workbook_path.write_bytes(response.content)
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["PO Items", "Parse Summary"]
