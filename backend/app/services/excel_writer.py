from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas import ParseResult

PO_ITEMS_HEADERS = [
    "Source File",
    "PO Number",
    "PO Date",
    "Ship To",
    "Item",
    "Material",
    "Description",
    "Manufacturer Part Number",
    "U/M",
    "Total Qty",
    "Qty Recd",
    "Qty Retd",
    "Unit Price",
    "Item Value",
    "Due Date",
    "Status / Warnings",
]

SUMMARY_HEADERS = [
    "Source File",
    "PO Number",
    "PO Date",
    "Status",
    "Line Item Count",
    "Warning Count",
    "Warnings",
    "Error",
]


def build_workbook(results: list[ParseResult]) -> bytes:
    workbook = Workbook()
    items_sheet = workbook.active
    items_sheet.title = "PO Items"
    summary_sheet = workbook.create_sheet("Parse Summary")

    _write_headers(items_sheet, PO_ITEMS_HEADERS)
    _write_headers(summary_sheet, SUMMARY_HEADERS)

    for result in results:
        _append_item_rows(items_sheet, result)
        _append_summary_row(summary_sheet, result)

    _format_sheet(items_sheet)
    _format_sheet(summary_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_headers(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _append_item_rows(sheet: Worksheet, result: ParseResult) -> None:
    if not result.line_items:
        return

    general_result_warnings = [warning for warning in result.warnings if not _is_item_specific_warning(warning)]
    result_warning_text = _join_messages(general_result_warnings)
    for item in result.line_items:
        status_parts = [result.status, result_warning_text, _join_messages(item.warnings)]
        sheet.append(
            [
                result.source_file,
                result.po_number,
                result.po_date,
                result.ship_to,
                item.item,
                item.material,
                item.description,
                item.manufacturer_part_number,
                item.uom,
                item.total_qty,
                item.qty_recd,
                item.qty_retd,
                item.unit_price,
                item.item_value,
                item.due_date,
                _join_messages([part for part in status_parts if part]),
            ]
        )


def _append_summary_row(sheet: Worksheet, result: ParseResult) -> None:
    warnings = _dedupe_messages(
        [*result.warnings, *[f"Item {item.item}: {warning}" for item in result.line_items for warning in item.warnings]]
    )

    sheet.append(
        [
            result.source_file,
            result.po_number,
            result.po_date,
            result.status,
            len(result.line_items),
            len(warnings),
            _join_messages(warnings),
            result.error or "",
        ]
    )


def _format_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(max(max_length + 2, 12), 60)
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _join_messages(messages: list[str]) -> str:
    return "; ".join(message for message in messages if message)


def _dedupe_messages(messages: list[str]) -> list[str]:
    deduped: list[str] = []
    for message in messages:
        if message and message not in deduped:
            deduped.append(message)
    return deduped


def _is_item_specific_warning(warning: str) -> bool:
    return warning.startswith("Item ") and ": " in warning
